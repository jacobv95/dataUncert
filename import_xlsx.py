import numpy as np
import openpyxl
import xlrd
import os.path
import re
import string
from fit import *
import matplotlib.pyplot as plt
try:
    from dataProcessing.variable import variable
except ModuleNotFoundError:
    from variable import variable


def readData(xlFile, dataStartCol, dataEndCol, uncertStartCol=None, uncertEndCol=None):
    dat = _readData(xlFile, dataStartCol, dataEndCol, uncertStartCol=uncertStartCol, uncertEndCol=uncertEndCol)
    return dat.dat


class _readData():

    def __init__(self, xlFile, dataStartCol, dataEndCol, uncertStartCol=None, uncertEndCol=None) -> None:

        # convert the coloumns
        self.dataStartCol = self.colToIndex(dataStartCol)
        self.dataEndCol = self.colToIndex(dataEndCol)
        self.uncertStartCol = self.colToIndex(uncertStartCol)
        self.uncertEndCol = self.colToIndex(uncertEndCol)

        # check the uncertanty range
        uncertCols = [self.uncertStartCol is None, self.uncertEndCol is None]
        if sum(uncertCols) not in [0, 2]:
            raise ValueError('You have provided one of the coloumn for the uncertanty but not the other')

        # check the number of coloumns
        nColsData = self.dataEndCol - self.dataStartCol + 1
        if not self.uncertStartCol is None:
            nColsUncert = self.uncertEndCol - self.uncertStartCol + 1
            if nColsData != nColsUncert:
                raise ValueError('The number of coloumns of the data is not equal to the number of coloumns for the uncertanty')
        self.nCols = nColsData

        # check the extension
        extension = os.path.splitext(xlFile)[1]
        supportedExtensions = ['.xls', '.xlsx']
        if extension not in supportedExtensions:
            raise ValueError(f'The file extension is not supported. The supported extension are {supportedExtensions}')

        # parse functions for the specific extension and get all sheets
        if extension == '.xls':
            self.wb = xlrd.open_workbook(xlFile)
            self.sheets = self.wb.sheets()

            def readCell(sheet, row, col):
                return sheet.cell(row, col).value

            def readRow(sheet, row):
                return [elem.value for elem in sheet.row(row)]

            def readCol(sheet, col):
                return [elem.value for elem in sheet.col(col)]

        elif extension == '.xlsx':
            self.wb = openpyxl.load_workbook(xlFile)
            self.sheets = [self.wb[elem] for elem in self.wb.sheetnames]

            def readCell(sheet, row, col):
                return sheet.cell(row + 1, col + 1).value

            def readRow(sheet, row):
                return [elem.value for elem in list(sheet.iter_rows())[row]]

            def readCol(sheet, col):
                return [elem.value for elem in list(sheet.iter_cols())[col]]

        self.readCell = readCell
        self.readRow = readRow
        self.readCol = readCol

        # read the data
        self.readData()

    def colToIndex(self, col):
        if col is None:
            return None
        if not isinstance(col, str):
            raise ValueError('The coloumn has to be a string')
        num = 0
        for c in col:
            if c in string.ascii_letters:
                num = num * 26 + (ord(c.upper()) - ord('A')) + 1
        return num

    def formatHeaders(self, header):
        out = []
        for head in header:

            # remove symbols and replace with _
            head = re.sub(r'[^\w]', '_', head.lower())

            # determine places with repeated "_"
            indexes_to_remove = []
            for i in range(len(head) - 1):
                if head[i] == "_":
                    if head[i + 1] == "_":
                        indexes_to_remove.append(i)

            # remove the indexes found in the previous step
            head = [char for char in head]
            for i in reversed(indexes_to_remove):
                head.pop(i)
            head = "".join(head)

            # add "_" to the begining of the name if the first letter is a digit
            if head[0].isnumeric():
                head = '_' + head

            # remove "_" if the last letter is "_"
            if head[-1] == "_" and len(head) != 1:
                head = head[0:-1]

            i, imax, done = 0, 100, False
            while not done and i <= imax:
                if i > 0:
                    h = head + f'_{i+2}'
                else:
                    h = head
                if h not in out:
                    out.append(h)
                    done = True

        return out

    def formatUnits(self, units):
        out = []
        for unit in units:
            if not unit is None and len(unit) > 0:
                # remove symbols and replace with _
                allowedCharacters = []
                allowedCharacters += list(string.ascii_letters)
                allowedCharacters += [str(num) for num in range(10)]
                allowedCharacters += ['/', '-']
                unit = list(unit)
                for i, char in enumerate(unit):
                    if char not in allowedCharacters:
                        unit.pop(i)
                unit = ''.join(unit)

                # remove "_" if the last letter is "_"
                if unit[-1] == "_" and len(unit) != 1:
                    unit = unit[0:-1]

                # remove "_" if the first letter is "_"
                if unit[0] == "_":
                    unit = unit[1:]

            out.append(unit)
        return out

    def readData(self):
        self.dat = _Data()

        for i, sheet in enumerate(self.sheets):

            sheetData = _Data(f's{i+1}')

            # determine the number of variables
            headers = self.readRow(sheet, 0)[0:self.nCols]
            headers = self.formatHeaders(headers)
            units = self.readRow(sheet, 1)[0:self.nCols]
            units = self.formatUnits(units)

            # determine the number of datapoints
            nDataPoints = []
            for i in range(self.nCols):
                nDataPoint = self.readCol(sheet, i)[2:]
                nDataPoint = sum([1 if elem not in ['', None] else 0 for elem in nDataPoint])
                nDataPoints.append(nDataPoint)
            if not all(elem == nDataPoints[0] for elem in nDataPoints):
                raise ValueError('There are not an equal amount of rows in the data')
            nDataPoint = nDataPoints[0]

            # read the data
            data = np.zeros([nDataPoint, self.nCols])
            for i in range(nDataPoint):
                for j in range(self.nCols):
                    data[i, j] = float(self.readCell(sheet, 2 + i, j))

            if not self.uncertStartCol is None:
                # determine the number of rows in the uncertanty
                nUncertanties = []
                for i in range(self.nCols):
                    nUncertanty = self.readCol(sheet, self.nCols + i)[2:]
                    nUncertanty = sum([1 if elem not in ['', None] else 0 for elem in nUncertanty])
                    nUncertanties.append(nUncertanty)
                if not all(elem == nUncertanties[0] for elem in nUncertanties):
                    raise ValueError('There are not an equal amount of rows in the uncertanty')
                nUncertanty = nUncertanties[0]

                # evaluate the number of rows of the uncertanty
                if nUncertanty not in [nDataPoint, nDataPoint * self.nCols]:
                    raise ValueError('The number of rows in the uncertanty has to be equal to the number of rows of data or equal to the number of rows of data multiplied with the number of coloumns in the data')

                if nUncertanty == nDataPoint:
                    # read the uncertanty
                    uncert = np.zeros([nDataPoint, self.nCols])
                    for i in range(nDataPoint):
                        for j in range(self.nCols):
                            uncert[i, j] = float(self.readCell(sheet, 2 + i, self.nCols + j))

                    # create the measurements uncertanties
                    for i in range(self.nCols):
                        name = headers[i]
                        unit = units[i]
                        val = np.array(data[:, i])
                        u = np.array(uncert[:, i])
                        var = variable(val, unit, uncert=u)

                        sheetData.addMeasurement(name, var)
                else:
                    # read the uncertanty
                    uncert = []
                    for i in range(nDataPoint):
                        u = np.zeros([self.nCols, self.nCols])
                        for j in range(self.nCols):
                            for k in range(self.nCols):
                                u[j, k] = float(self.readCell(sheet, 2 + i * self.nCols + j, self.nCols + k))
                        uncert.append(u)

                    # create the measurements with covariance uncertanties
                    vars = []
                    for i in range(self.nCols):
                        name = headers[i]
                        unit = units[i]
                        val = np.array(data[:, i])
                        u = np.array([elem[i, i] for elem in uncert])
                        var = variable(val, unit, uncert=u)
                        vars.append(var)

                    for i in range(self.nCols):
                        covariance = [elem[:, i] for elem in uncert]
                        for j in range(self.nCols):
                            if i != j:
                                cov = [elem[j] for elem in covariance]
                                vars[i]._addCovariance(vars[j], cov)

                    for head, var in zip(headers, vars):
                        sheetData.addMeasurement(head, var)
            else:
                # create the measurements without uncertanties
                for i in range(self.nCols):
                    name = headers[i]
                    unit = units[i]
                    val = np.array(data[:, i])
                    var = variable(val, unit)
                    sheetData.addMeasurement(name, var)

            self.dat.addSheet(sheetData.name, sheetData)


class _Data():
    def __init__(self, name='') -> None:
        self.name = name
        self.measurements = []
        self.measurementNames = []
        self.sheets = []

    def addSheet(self, name, sheet):
        self.sheets.append(sheet)
        setattr(self, name, sheet)

    def addMeasurement(self, name, var):
        self.measurements.append(var)
        self.measurementNames.append(name)
        setattr(self, name, var)

    def printContents(self, suffix=None):
        for sheet in self.sheets:
            sheet.printContents(self.name)
            print('')

        for name in self.measurementNames:
            if suffix is None:
                print(f'{self.name}.{name}')
            else:
                print(f'{suffix}.{self.name}.{name}')


def main():
    a = readData('a.xls', 'A', 'K', 'L', 'V')
    b = readData('b.xlsx', 'A', 'B')
    c = readData('c.xlsx', 'A', 'B', 'C', 'D')
    d = readData('d.xlsx', 'A', 'B', 'C', 'D')
    a.printContents()
    b.printContents()
    c.printContents()
    d.printContents()

    F = lin_fit(a.s1.t_phe_in_oil, a.s1.t_phe_out_glycol, p0=[1, 50])
    fig, ax = plt.subplots()
    F.plot(ax)
    F.scatter(ax)
    ax.legend()
    ax.set_xlabel('temp1')
    ax.set_ylabel('temp2')
    F.addUnitToLabels(ax)
    plt.show()


if __name__ == '__main__':
    main()
